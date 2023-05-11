#!usr/bin/env

# import libraries
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont


import openpyxl
import cv2

#-----------------------------------------------------------------------------
import tkinter as tk
from tkinter import messagebox
from tkinter import font

mouse_position = None
class App:
    def __init__(self, master):
        self.master = master
        self.master.title("Invitation Namer")

        bold_font = font.Font(weight='bold', size=10)
        
        # Create the text fields
        tk.Label(master, text="Excel file path").grid(row=0)
        tk.Label(master, text="Invitation path").grid(row=1)
        tk.Label(master, text="Height").grid(row=2)
        tk.Label(master, text="Width").grid(row=3)
        tk.Label(master, text="Font").grid(row=4)
        tk.Label(master, text="Font size").grid(row=5)
        tk.Label(master, text="Font color").grid(row=6)
        tk.Label(master, text="keep text field empty to get the default settings", font=bold_font, anchor='w').grid(row=7)
        tk.Label(master, text="Enter c in height or width for center alignment", font=bold_font, anchor='w').grid(row=8)
        tk.Label(master, text="Press create button and right click on image to select the writing location\n and close the image window", font=bold_font, anchor='w').grid(row=9)
        
        self.field1 = tk.Entry(master)
        self.field2 = tk.Entry(master)
        self.field3 = tk.Entry(master)
        self.field4 = tk.Entry(master)
        self.field5 = tk.Entry(master)
        self.field6 = tk.Entry(master)
        self.field7 = tk.Entry(master)
        self.field1.grid(row=0, column=1)
        self.field2.grid(row=1, column=1)
        self.field3.grid(row=2, column=1)
        self.field4.grid(row=3, column=1)
        self.field5.grid(row=4, column=1)
        self.field6.grid(row=5, column=1)
        self.field7.grid(row=6, column=1)

        # Create the create button
        tk.Button(master, text="Create", command=self.create, bg="red", fg="white").grid(row=7, column=1)

    def create(self):
        # Get the values from the text fields
        value1 = self.field1.get()
        value2 = self.field2.get()
        value3 = self.field3.get()
        value4 = self.field4.get()
        value5 = self.field5.get()
        value6 = self.field6.get()
        value7 = self.field7.get()

        # Print the values to the console
        # print("Excel path:", value1)
        # print("Invitation path:", value2)
        # print("Height:", value3)
        # print("Width:", value4)
        # print("Font:", value5)
        # print("Font size:", value6)
        # print("Font color:", value7)
        
        if value1 == '':
            value1 = "name_list.xlsx"
        if value2 == '':
            value2 = "final_invitations/final_invitation.jpg"
        # Open the Workbook
        wb = openpyxl.load_workbook(value1)
        
        # Open the worksheet
        
        ws = wb['Sheet1']
        
        img = Image.open(value2)

        # Load the image
        img2 = cv2.imread(value2)
        # Display the image
        cv2.imshow("Image", img2)

        # Set the mouse callback function for the window
        cv2.setMouseCallback("Image", get_pixel)

        # Wait for a key press and then close the window
        cv2.waitKey(0)
        cv2.destroyAllWindows()

        # draw image object
        I1 = ImageDraw.Draw(img)

        width = img.width
        height = img.height
        
        # display width and height
        # print("The height of the image is: ", height)
        # print("The width of the image is: ", width)
        
        
            
            
        if value5 == '':
            value5 = "Fonts/Demo_ConeriaScript.ttf"
        if value6 == '':
            value6 = 35
        if value7 == '':
            value7 = (0,0,0)
        else:
            value7 = eval(value7)
        # Iterate over rows and columns until the last entry
        # Iterate over rows and columns until the last entry
        for i in range(ws.max_row+1):
            img = Image.open(value2)
            # print("i = ",i)
            
            # Print the value of the current cell
            try:
                only_name = ws.cell(row=i+1, column=2).value
                # print("only name = ", only_name)
                name = ws.cell(row=i+1, column=1).value + " " + only_name
                # print(name)
            except:
                break

            # Draw image object
            I1 = ImageDraw.Draw(img)
            myFont = ImageFont.truetype(value5, int(value6))
            # Get the width and height of the text
            _, _, w, h = I1.textbbox((0, 0), name, font=myFont)

            # Center the text on the image
            x = (width-w)/2
            y = (height-h)/2
            
            if value3 == '':
                y = mouse_position[1]
            if value4 == '':
                x = mouse_position[0]

            # Add Text to an image
            # print("checking mouse positions:", mouse_position)
            I1.text((x, y), name , font=myFont, fill =value7)

            # save image
            img.save("images/"+only_name+".png")

        # Show a success message in a pop-up window
        messagebox.showinfo("Success", "Invitations are created successfully!")

# Create a mouse callback function
def get_pixel(event, x, y, flags, param):
    if event == cv2.EVENT_RBUTTONDOWN:
        # print("Pixel location: ({}, {})".format(y, x))
        global mouse_position
        mouse_position = (x, y)

root = tk.Tk()
app = App(root)
root.mainloop()
